from datetime import date as date_type
from io import BytesIO

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select

from app.api.deps import CurrentUser, DbDep
from app.models.category import Category
from app.models.debt import Debt
from app.models.enums import DebtDirection, OperationType
from app.models.transaction import Transaction
from app.schemas.transaction import TransactionCreate, TransactionOut, TransactionUpdate

router = APIRouter(prefix="/transactions", tags=["transactions"])


def _check_category(db, current_user, category_id):
    if category_id is None:
        return
    cat = db.get(Category, category_id)
    if not cat or cat.user_id != current_user.id:
        raise HTTPException(status_code=400, detail="Invalid category_id")


def _get_owned(db, current_user, tx_id) -> Transaction:
    tx = db.get(Transaction, tx_id)
    if not tx or tx.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return tx


@router.get("", response_model=list[TransactionOut])
def list_transactions(
    db: DbDep,
    current_user: CurrentUser,
    type: OperationType | None = None,
    date_from: date_type | None = None,
    date_to: date_type | None = None,
):
    stmt = select(Transaction).where(Transaction.user_id == current_user.id)
    if type is not None:
        stmt = stmt.where(Transaction.type == type)
    if date_from is not None:
        stmt = stmt.where(Transaction.date >= date_from)
    if date_to is not None:
        stmt = stmt.where(Transaction.date <= date_to)
    stmt = stmt.order_by(Transaction.date.desc(), Transaction.id.desc())
    return db.scalars(stmt).all()


# --- палитра и формат для Excel-выгрузки ---
_GREEN = "2E7D32"        # доход / мне должны
_RED = "C62828"          # расход / я должен
_HEADER_FILL = "1F4E78"  # шапка таблиц
_GREEN_FILL = "E8F5E9"   # светлая подложка доходов
_RED_FILL = "FDECEA"     # светлая подложка расходов
_MONEY_FMT = '#,##0.00" ₽"'


@router.get("/export")
def export_transactions(
    db: DbDep,
    current_user: CurrentUser,
    type: OperationType | None = None,
):
    """Выгрузка операций в Excel-файл (.xlsx). Используется и сайтом, и ботом.

    Слева — таблица операций с цветовой разметкой (доход — зелёный,
    расход — красный) и итогами внизу. Справа — отдельная таблица долгов.
    """
    stmt = select(Transaction).where(Transaction.user_id == current_user.id)
    if type is not None:
        stmt = stmt.where(Transaction.type == type)
    stmt = stmt.order_by(Transaction.date.desc(), Transaction.id.desc())
    txs = db.scalars(stmt).all()

    cats = {
        c.id: c.name
        for c in db.scalars(select(Category).where(Category.user_id == current_user.id)).all()
    }
    debts = db.scalars(select(Debt).where(Debt.user_id == current_user.id)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Операции"

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    green_font = Font(color=_GREEN, bold=True)
    red_font = Font(color=_RED, bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    def style_header(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    # ===== основная таблица: операции (колонки A–E) =====
    for col, h in enumerate(["Дата", "Тип", "Категория", "Сумма", "Описание"], start=1):
        style_header(ws.cell(row=1, column=col, value=h))

    total_income = 0.0
    total_expense = 0.0
    row = 2
    for t in txs:
        is_income = t.type == OperationType.income
        amount = float(t.amount)
        if is_income:
            total_income += amount
        else:
            total_expense += amount
        money_font = green_font if is_income else red_font
        fill = PatternFill("solid", fgColor=_GREEN_FILL if is_income else _RED_FILL)

        ws.cell(row=row, column=1, value=t.date.isoformat()).border = border
        type_cell = ws.cell(row=row, column=2, value="Доход" if is_income else "Расход")
        type_cell.font = money_font
        type_cell.fill = fill
        type_cell.border = border
        ws.cell(row=row, column=3, value=cats.get(t.category_id, "")).border = border
        amt_cell = ws.cell(row=row, column=4, value=amount)
        amt_cell.number_format = _MONEY_FMT
        amt_cell.font = money_font
        amt_cell.border = border
        ws.cell(row=row, column=5, value=t.description or "").border = border
        row += 1

    # итоги под операциями
    row += 1
    ws.cell(row=row, column=3, value="Итого доходы:").font = bold
    inc = ws.cell(row=row, column=4, value=total_income)
    inc.number_format = _MONEY_FMT
    inc.font = green_font
    row += 1
    ws.cell(row=row, column=3, value="Итого расходы:").font = bold
    exp = ws.cell(row=row, column=4, value=total_expense)
    exp.number_format = _MONEY_FMT
    exp.font = red_font
    row += 1
    balance = total_income - total_expense
    ws.cell(row=row, column=3, value="Баланс:").font = bold
    bal = ws.cell(row=row, column=4, value=balance)
    bal.number_format = _MONEY_FMT
    bal.font = Font(bold=True, color=_GREEN if balance >= 0 else _RED)

    # ===== таблица долгов справа (колонки G–J; F — зазор) =====
    for i, h in enumerate(["Контрагент", "Кто кому", "Сумма", "Статус"]):
        style_header(ws.cell(row=1, column=7 + i, value=h))

    if debts:
        owe_total = 0.0
        owed_total = 0.0
        drow = 2
        for d in debts:
            i_owe = d.direction == DebtDirection.i_owe
            amount = float(d.amount)
            if not d.is_settled:
                if i_owe:
                    owe_total += amount
                else:
                    owed_total += amount
            money_font = red_font if i_owe else green_font
            fill = PatternFill("solid", fgColor=_RED_FILL if i_owe else _GREEN_FILL)

            ws.cell(row=drow, column=7, value=d.counterparty).border = border
            dir_cell = ws.cell(row=drow, column=8, value="Я должен" if i_owe else "Мне должны")
            dir_cell.font = money_font
            dir_cell.fill = fill
            dir_cell.border = border
            amt_cell = ws.cell(row=drow, column=9, value=amount)
            amt_cell.number_format = _MONEY_FMT
            amt_cell.font = money_font
            amt_cell.border = border
            ws.cell(row=drow, column=10, value="Погашен" if d.is_settled else "Активен").border = border
            drow += 1

        drow += 1
        ws.cell(row=drow, column=8, value="Я должен:").font = bold
        o1 = ws.cell(row=drow, column=9, value=owe_total)
        o1.number_format = _MONEY_FMT
        o1.font = red_font
        drow += 1
        ws.cell(row=drow, column=8, value="Мне должны:").font = bold
        o2 = ws.cell(row=drow, column=9, value=owed_total)
        o2.number_format = _MONEY_FMT
        o2.font = green_font
    else:
        ws.cell(row=2, column=7, value="Долгов пока нет.")

    # ширина колонок
    widths = {"A": 12, "B": 10, "C": 18, "D": 16, "E": 28, "F": 3,
              "G": 18, "H": 13, "I": 16, "J": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="operations.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("", response_model=TransactionOut, status_code=201)
def create_transaction(data: TransactionCreate, db: DbDep, current_user: CurrentUser):
    _check_category(db, current_user, data.category_id)
    tx = Transaction(
        user_id=current_user.id,
        amount=data.amount,
        type=data.type,
        category_id=data.category_id,
        description=data.description,
        date=data.date or date_type.today(),
    )
    db.add(tx)
    db.commit()
    db.refresh(tx)
    return tx


@router.patch("/{tx_id}", response_model=TransactionOut)
def update_transaction(tx_id: int, data: TransactionUpdate, db: DbDep, current_user: CurrentUser):
    tx = _get_owned(db, current_user, tx_id)
    payload = data.model_dump(exclude_unset=True)
    if "category_id" in payload:
        _check_category(db, current_user, payload["category_id"])
    for field, val in payload.items():
        setattr(tx, field, val)
    db.commit()
    db.refresh(tx)
    return tx


@router.delete("/{tx_id}", status_code=204)
def delete_transaction(tx_id: int, db: DbDep, current_user: CurrentUser):
    tx = _get_owned(db, current_user, tx_id)
    db.delete(tx)
    db.commit()
